from fastapi import FastAPI, File, UploadFile
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import re
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO

stream = None

app = FastAPI()


app.mount("/static", StaticFiles(directory="client/build/static"), name="static")


def find_column(ws, header_row, regexp):
    for row in ws.iter_rows(min_row=header_row, max_row=header_row):
        for cell in row:
            if re.match(regexp, str(cell.value), re.IGNORECASE):
                column = cell.column
                break
    return column


def highlight_row(ws, row_number):
    yellow_fill = PatternFill(
        fill_type="solid", start_color="FFFF00", end_color="FFFF00"
    )
    for cell in ws[row_number]:
        cell.fill = yellow_fill


@app.get("/")
async def read_index():
    return FileResponse("client/build/index.html")


@app.get("/favicon.ico")
async def favicon():
    return FileResponse("client/build/favicon.ico")


@app.post("/upload/")
async def create_upload_file(file: UploadFile = File(...)):
    global stream
    contents = await file.read()
    wb = load_workbook(filename=BytesIO(contents))
    ws = wb.active

    # Find header row
    for row in ws.iter_rows():
        for cell in row:
            if re.match(
                r"(^.*part number.*$)|(^.*item name.*$)", str(cell.value), re.IGNORECASE
            ):
                header_row = cell.row  # save header row
                partnumber_col = cell.column  # save partnumber column
                break

    # Find key columns
    description_col = find_column(ws, header_row, regexp=r"^.*Description.*$")
    price_col = find_column(ws, header_row, regexp=r"^.*list\s?price.*$")
    partnumber_col = find_column(
        ws, header_row, regexp=r"(^.*part number.*$)|(^.*item name.*$)"
    )
    leadtime_col = find_column(ws, header_row, regexp=r"^.*lead\s?time.*$")
    service_col = find_column(ws, header_row, regexp=r"^.*service\s?duration.*$")

    print(f"Header row: {header_row}")
    print(f"PN Column: {partnumber_col}")
    print(f"Lead Time Column: {leadtime_col}")
    print(f"List Price Column: {price_col}")
    print(f"Service Duration Column: {service_col}")

    # Mark zero VAT items
    for row_number in range(header_row + 1, ws.max_row + 1):

        # Gather key values
        partnumber = ws.cell(row=row_number, column=partnumber_col).value
        description = ws.cell(row=row_number, column=description_col).value
        price = ws.cell(row=row_number, column=price_col).value
        lead_time = ws.cell(row=row_number, column=leadtime_col).value
        service_duration = ws.cell(row=row_number, column=service_col).value

        # Normalize lead time (it may be NoneType or str)
        if isinstance(lead_time, str):
            try:
                lead_time = int(lead_time)
            # Handle non-numerical string, e.g. "N/A"
            except ValueError:
                lead_time = None
        # Normalize service duration (it may be int or str)
        if isinstance(service_duration, str):
            try:
                service_duration = int(service_duration)
            # Handle non-numerical string, e.g. "---"
            except ValueError:
                service_duration = None
        # Exclude partnumbers
        if not re.match(r"(^CON-.*)|(^SG3.*)", str(partnumber)):
            # Check partnumber and price
            if (
                re.match(
                    r"(^[LRS]-)|(^LIC-)|(.*[1-9]Y$)|(.*-[1-9]Y-.*)", str(partnumber)
                )
                and int(price) > 0
            ):
                highlight_row(ws, row_number)
            # Check description and price
            elif (
                re.match(r".*e-?[\s-]?delivery.*", str(description), re.IGNORECASE)
                and int(price) > 0
            ):
                highlight_row(ws, row_number)
            # Check lead time
            elif lead_time and 0 < lead_time <= 10 and int(price) > 0:
                highlight_row(ws, row_number)
            # Check service duration
            elif service_duration and int(price) > 0:
                highlight_row(ws, row_number)

    # Save resulting sheet in temp file
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

    return {"status": "ok"}


@app.get("/download/")
async def download_file():
    global stream
    if stream:
        response = StreamingResponse(
            BytesIO(stream),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["Content-Disposition"] = "attachment; filename=estimate.xlsx"
    else:
        response = {"message": "nothing to download"}
    return response
